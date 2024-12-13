import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

# Function to extract financial data from PDF
def extract_financial_data_ai(pdf_path):
    financial_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                for line in lines:
                    # Check for keywords commonly associated with financial data
                    if any(keyword in line for keyword in ["Scrutiny fee", "Centage Charges", "Development Charges", "OSR Fee", "Extent", "No. of Plots"]):
                        financial_data.append(line)
    return financial_data

# Function to process extracted data
def process_financial_data_ai(data):
    processed_data = {}
    for line in data:
        if "Scrutiny fee" in line:
            processed_data['Scrutiny Fee'] = float(line.split(':')[-1].strip().replace(',', '').replace('Rs.', '').strip())
        elif "Centage Charges" in line:
            processed_data['Centage Charges'] = float(line.split(':')[-1].strip().replace(',', '').replace('Rs.', '').strip())
        elif "Development Charges" in line:
            processed_data['Development Charges'] = float(line.split(':')[-1].strip().replace(',', '').replace('Rs.', '').strip())
        elif "OSR Fee" in line:
            processed_data['OSR Fee'] = float(line.split(':')[-1].strip().replace(',', '').replace('Rs.', '').strip())
        elif "Extent" in line:
            processed_data['Site Extent'] = line.split(':')[-1].strip()
        elif "No. of Plots" in line:
            processed_data['No. of Plots'] = int(line.split(':')[-1].strip())
    return processed_data

# Function to save processed data to Excel
def save_to_excel_ai(processed_data, excel_path):
    df = pd.DataFrame([processed_data])

    if not os.path.exists(excel_path):
        df.to_excel(excel_path, index=False, sheet_name='Financial Data')
    else:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
            book = load_workbook(excel_path)
            writer.book = book
            if 'Financial Data' not in book.sheetnames:
                df.to_excel(writer, index=False, sheet_name='Financial Data')
            else:
                existing_df = pd.read_excel(excel_path, sheet_name='Financial Data')
                updated_df = pd.concat([existing_df, df], ignore_index=True)
                updated_df.to_excel(writer, index=False, sheet_name='Financial Data')

# Main script to process multiple PDFs
def main():
    pdf_paths = input("Enter the paths to the PDF files (comma-separated): ").strip().split(',')
    excel_path = input("Enter the path to the Excel file (or desired output path): ").strip()

    for pdf_path in pdf_paths:
        pdf_path = pdf_path.strip()
        print(f"Processing {pdf_path}...")
        financial_data = extract_financial_data_ai(pdf_path)
        processed_data = process_financial_data_ai(financial_data)
        save_to_excel_ai(processed_data, excel_path)

    print("Financial analysis complete. Data saved to Excel.")

if __name__ == "__main__":
    main()
